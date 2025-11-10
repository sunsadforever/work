import random
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
import os
service = Service(r"C:\Users\jerry\桌面\chromedriver-win64\chromedriver.exe")
last_name=None
last_name_type=input('你想要姓氏隨機還是指定: ')
if last_name_type=='指定':
    last_name_type='custom'
    last_name=input('請輸入你想要的姓氏: ')
else:
    last_name_type='random'

gender=input('請輸入你想要的性別，男生、女生或任何: ')
if gender=='男生':
    gender='B'
elif gender=='女生':
    gender='G'
else:
    gender='any'
popularity=input('你想要什麼熱門程度(有任意、熱門、一般、罕見): ')
if popularity=='熱門':
    popularity='popular'
elif popularity=='一般':
    popularity='familiar'
elif popularity=='罕見':
    popularity='unusual'
else:
    popularity='any'
driver = webdriver.Chrome(service=service)

if last_name is None:
    url=f'https://www.namechef.co/zh/name-generator/chinese/?last_name_type={last_name_type}&gender={gender}&popularity%5B%5D={popularity}'
else:
    url=f'https://www.namechef.co/zh/name-generator/chinese/?last_name_type={last_name_type}&last_name={last_name}&gender={gender}&popularity%5B%5D={popularity}'
driver.get(url)
name_data=[]
for index in range(1,21): #有20個姓名
    name=driver.find_element(By.ID,f'name_{index}').text
    name_data.append(name)
real_name=random.choice(name_data)
print(real_name)

#chinese-中文
# last_name_type-姓氏，random是隨機，custom是指定，若有custom，那後面會接last_name，後面則是中文姓氏
# gender是性別 有any、B、G可選
# popularity%5B%5D是熱門程度 有any、popular、familiar、unusual可選
ws = load_workbook(r"名字.xlsx")
wb=ws.active
row=2
if wb['B1'].value is None:
    wb['B1'] = '名字'
while True:
    if wb[f'A{row}'].value is None:
        wb[f'A{row}']=f'角色{row-1}'
    if wb[f'B{row}'].value is None:
        wb[f'B{row}']=real_name
        break
    row+=1
ws.save(r"名字.xlsx")
os.startfile(r"名字.xlsx")
